#Doc is called a workbook
#Sinlge workbook called 
#Each file can contain multiples sheets
#Columns and rows
#Cells intersection of columns and rows

import openpyxl, os

os.chdir('c:\\users\\nicho\\Desktop\\PythonPrograms\\TestFiles')
#Change directory to where the excel file is

workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook))

print(workbook.get_sheet_names())
sheet = workbook.get_sheet_by_name('Sheet1')
#Returns sheet object
print(sheet)
cell = sheet['A1']
#Returns the cell object

print(sheet['A1'].value)
print(cell.value)
#evaluates to cell value

print(sheet['C1'].value)
#Column formatted as all numbers, return an int

print(sheet.cell(row=1, column =2))
#returns cell object, same as sheet['B1']
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
