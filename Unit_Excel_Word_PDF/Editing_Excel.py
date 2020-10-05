import openpyxl, os

os.chdir('c:\\users\\nicho\\Desktop\\PythonPrograms\\TestFiles')
#Change directory

wb = openpyxl.Workbook()
#Creates a new workbook object
#Starts with a single sheet named sheet

print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value)
#Evaluates to None, blank worksheet

sheet2 = wb.create_sheet()
sheet['A1'] = 42
sheet['A2'] = 'Hello'

print(wb.get_sheet_names())
sheet2.title = 'My New sheet name'
print(wb.get_sheet_names())

wb.create_sheet(index=0, title='My Other Sheet')
 #Makes this sheet the first sheet, and provides title
print(wb.get_sheet_names())

wb.save('example2.xlsx')
    #Saves workbook in current directory
    #When editing, save with new name, do not override old file



openpyxl.load_workbook('example2.xlsx')
#Can edit now
